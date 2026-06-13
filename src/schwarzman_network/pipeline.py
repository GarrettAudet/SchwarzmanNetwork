from __future__ import annotations

import csv
import json
import tempfile
import time
from pathlib import Path

from openpyxl import load_workbook

from .config import AUDIT_DIR, DATA_DIR, INTERIM_DIR, PROCESSED_DIR, PUBLIC_DIR, RAW_DIR, SEED_DIR, ensure_data_dirs
from .enrichment.company import enrich_company
from .enrichment.enrichlayer import (
    ENRICHLAYER_FIELDNAMES,
    EnrichLayerClient,
    EnrichLayerError,
    normalize_enrichlayer_record,
)
from .enrichment.linkedin_api import BrightDataLinkedInClient
from .enrichment.schema import normalize_brightdata_record
from .matching.adjudicator import candidate_evidence_json, choose_linkedin_candidate
from .matching.linkedin_quality import review_linkedin_quality
from .matching.merge import append_new_official_scholars
from .models import EmploymentObservation, LinkedInProfile, Scholar, clean_text, scholar_key, utc_now_iso
from .official.cohort import graduation_year_from_cohort
from .official.extractor import write_official_snapshot
from .official.parser import normalize_official_json
from .search.collector import collect_linkedin_candidates
from .search.normalize import is_linkedin_profile_url, linkedin_slug, normalize_linkedin_url
from .storage.build_database import build_database
from .storage.export_public import export_public


def _write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            "w",
            encoding="utf-8",
            newline="",
            dir=path.parent,
            prefix=f".{path.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            tmp_path = Path(handle.name)
            writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
            writer.writeheader()
            writer.writerows(rows)
        for attempt in range(8):
            try:
                tmp_path.replace(path)
                break
            except PermissionError:
                if attempt == 7:
                    raise
                time.sleep(0.5 * (attempt + 1))
    finally:
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
    return path


def _blank_if_na(value: object) -> str:
    text = clean_text(value)
    return "" if text.lower() in {"n/a", "na", "none", "null", "-"} else text


def _console_safe(value: object) -> str:
    return clean_text(value).encode("ascii", "backslashreplace").decode("ascii")


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _record_input_url(record: dict[str, object], fallback_url: str = "") -> str:
    input_value = record.get("input")
    nested_url = input_value.get("url") if isinstance(input_value, dict) else ""
    return normalize_linkedin_url(clean_text(record.get("input_url") or nested_url or record.get("url") or fallback_url))


def _iter_brightdata_cached_records(raw_path: Path) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    if raw_path.exists():
        for line in raw_path.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            item = json.loads(line)
            record = item.get("record", {})
            if isinstance(record, dict):
                items.append(
                    {
                        "input_url": _record_input_url(record, clean_text(item.get("input_url"))),
                        "fetched_at": clean_text(item.get("fetched_at")),
                        "record": record,
                    }
                )

    for path in [DATA_DIR / "brightdata_profile_enrichment.json", DATA_DIR / "brightdata_profile_enrichment_pilot.json"]:
        if not path.exists():
            continue
        payload = json.loads(path.read_text(encoding="utf-8"))
        for item in payload.get("records", []):
            if not isinstance(item, dict):
                continue
            record = item.get("record", {})
            if not isinstance(record, dict):
                continue
            items.append(
                {
                    "input_url": _record_input_url(record, clean_text(item.get("inputUrl") or item.get("input_url"))),
                    "fetched_at": clean_text(item.get("fetchedAt") or item.get("fetched_at") or payload.get("generatedAt")),
                    "record": record,
                }
            )
    normalized_path = AUDIT_DIR / "brightdata_profile_decisions.csv"
    if normalized_path.exists():
        for row in _read_csv(normalized_path):
            input_url = normalize_linkedin_url(row.get("input_url", ""))
            if input_url == "N/A":
                continue
            items.append(
                {
                    "input_url": input_url,
                    "fetched_at": clean_text(row.get("fetched_at")),
                    "record": {
                        "input_url": input_url,
                        "name": row.get("name", ""),
                        "position": row.get("position", ""),
                        "current_company_name": row.get("current_company_name", ""),
                        "current_company_company_id": row.get("current_company_company_id", ""),
                        "profile_city": row.get("profile_city", ""),
                        "profile_location": row.get("profile_location", ""),
                        "job_location": row.get("job_location", ""),
                        "job_location_source": row.get("job_location_source", ""),
                        "city": row.get("city", ""),
                        "location": row.get("location", ""),
                    },
                }
            )
    return items


def _official_lookup_by_name() -> dict[str, dict[str, object]]:
    paths = [RAW_DIR / "official_scholars.json", DATA_DIR / "official_schwarzman_all.json"]
    for path in paths:
        if not path.exists():
            continue
        payload = json.loads(path.read_text(encoding="utf-8"))
        rows = normalize_official_json(payload)
        if rows:
            return {clean_text(row.get("name")).lower(): row for row in rows}
    return {}


def import_workbook(input_path: Path, seed_dir: Path = SEED_DIR, sheet_name: str = "LinkedIn Addresses") -> dict[str, object]:
    ensure_data_dirs()
    wb = load_workbook(input_path, read_only=True, data_only=True)
    sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows_iter)]
    header_index = {header: index for index, header in enumerate(headers)}

    def value(row: tuple[object, ...], header: str) -> str:
        index = header_index.get(header)
        return clean_text(row[index]) if index is not None and index < len(row) else ""

    official_by_name = _official_lookup_by_name()
    seen_ids: set[str] = set()
    scholars: list[dict[str, object]] = []
    linkedin_rows: list[dict[str, object]] = []
    observation_rows: list[dict[str, object]] = []
    imported_at = utc_now_iso()

    for row in rows_iter:
        name = value(row, "Scholar Name")
        cohort = value(row, "Cohort").upper()
        official = official_by_name.get(name.lower())
        if name and not cohort and official:
            cohort = clean_text(official.get("cohort")).upper()
        if not name or not cohort:
            continue
        base_id = scholar_key(name, cohort)
        scholar_id = base_id
        suffix = 2
        while scholar_id in seen_ids:
            scholar_id = f"{base_id}-{suffix}"
            suffix += 1
        seen_ids.add(scholar_id)

        country = _blank_if_na(value(row, "Country")) or clean_text(official.get("country") if official else "")
        linkedin_url = normalize_linkedin_url(value(row, "LinkedIn"))
        location = _blank_if_na(value(row, "Location"))
        company = _blank_if_na(value(row, "Company"))
        job = _blank_if_na(value(row, "Job"))
        scholar = Scholar(
            scholar_id=scholar_id,
            scholar_name=name,
            cohort=cohort,
            country=country,
            graduation_year=graduation_year_from_cohort(cohort),
            official_url=clean_text(official.get("sourceUrl") if official else ""),
            official_bio=clean_text(official.get("bio") if official else ""),
            source="workbook_seed",
        )
        profile = LinkedInProfile(
            scholar_id=scholar_id,
            scholar_name=name,
            cohort=cohort,
            linkedin_url=linkedin_url,
            linkedin_slug=linkedin_slug(linkedin_url),
            status="present" if linkedin_url != "N/A" else "missing",
            source="workbook_seed",
        )
        observation = EmploymentObservation(
            scholar_id=scholar_id,
            scholar_name=name,
            cohort=cohort,
            observed_at=imported_at,
            current_location=location,
            current_company=company,
            current_title=job,
            source_kind="workbook_bootstrap",
            source_url=linkedin_url if linkedin_url != "N/A" else "",
            confidence="manual_workbook",
            raw_source=input_path.name,
        )
        scholars.append(scholar.row())
        linkedin_rows.append(profile.row())
        observation_rows.append(observation.row())

    _write_csv(seed_dir / "scholars.csv", scholars, list(Scholar("", "", "").row().keys()))
    _write_csv(seed_dir / "linkedin_profiles.csv", linkedin_rows, list(LinkedInProfile("", "", "", "", "", "", "").row().keys()))
    _write_csv(
        seed_dir / "employment_observations.csv",
        observation_rows,
        list(EmploymentObservation("", "", "", "", "", "", "", "", "", "", "").row().keys()),
    )
    return {
        "sheet": sheet.title,
        "scholars": len(scholars),
        "linkedin_present": sum(1 for row in linkedin_rows if row["linkedin_url"] != "N/A"),
        "linkedin_missing": sum(1 for row in linkedin_rows if row["linkedin_url"] == "N/A"),
        "seed_dir": str(seed_dir),
    }


def fetch_official(output_path: Path | None = None) -> Path:
    ensure_data_dirs()
    return write_official_snapshot(output_path or RAW_DIR / "official_scholars.json")


def sync_official(official_path: Path | None = None, seed_dir: Path = SEED_DIR) -> dict[str, int]:
    path = official_path or RAW_DIR / "official_scholars.json"
    return append_new_official_scholars(seed_dir, path)


def find_missing_linkedin(
    seed_dir: Path = SEED_DIR,
    limit: int = 0,
    providers: list[str] | None = None,
    matching_mode: str = "llm",
) -> dict[str, object]:
    ensure_data_dirs()
    scholars = {row["scholar_id"]: row for row in _read_csv(seed_dir / "scholars.csv")}
    linkedin_rows = _read_csv(seed_dir / "linkedin_profiles.csv")
    pending = [row for row in linkedin_rows if not row.get("linkedin_url") or row.get("linkedin_url") == "N/A"]
    if limit > 0:
        pending = pending[:limit]

    provider_names = providers or ["google", "brave", "bing"]
    decisions: list[dict[str, object]] = []
    accepted = 0
    for row in pending:
        scholar = scholars.get(row["scholar_id"], {})
        name = scholar.get("scholar_name", row.get("scholar_name", ""))
        country = scholar.get("country", "")
        cohort = scholar.get("cohort", row.get("cohort", ""))
        candidates = collect_linkedin_candidates(
            name,
            country,
            cohort,
            providers=provider_names,
        )
        url, decision = choose_linkedin_candidate(
            name,
            candidates,
            country=country,
            cohort=cohort,
            official_bio=scholar.get("official_bio", ""),
            mode=matching_mode,
        )
        decisions.append(
            {
                "scholar_id": row["scholar_id"],
                "scholar_name": row.get("scholar_name", ""),
                "cohort": row.get("cohort", ""),
                "accepted_url": url,
                "accepted": str(bool(decision.get("accepted"))),
                "adjudicator": clean_text(decision.get("adjudicator")),
                "confidence": clean_text(decision.get("confidence")),
                "reason": clean_text(decision.get("reason")),
                "rationale": clean_text(decision.get("rationale")),
                "selected_candidate_id": clean_text(decision.get("selected_candidate_id")),
                "selected_candidate_url": clean_text(decision.get("selected_candidate_url")),
                "candidate_count": len(candidates),
                "candidate_evidence": candidate_evidence_json(candidates),
                "providers": ";".join(provider_names),
                "matching_mode": matching_mode,
                "checked_at": utc_now_iso(),
            }
        )
        if url != "N/A" and decision.get("accepted"):
            row["linkedin_url"] = url
            row["linkedin_slug"] = linkedin_slug(url)
            row["status"] = "present"
            row["source"] = "search_llm_verified" if decision.get("adjudicator") == "llm" else "search_heuristic_verified"
            accepted += 1

    if linkedin_rows:
        _write_csv(seed_dir / "linkedin_profiles.csv", linkedin_rows, list(linkedin_rows[0].keys()))
    if decisions:
        _write_csv(AUDIT_DIR / "linkedin_search_decisions.csv", decisions, list(decisions[0].keys()))
    return {
        "checked": len(pending),
        "accepted": accepted,
        "matching_mode": matching_mode,
        "audit": str(AUDIT_DIR / "linkedin_search_decisions.csv"),
    }


def enrich_brightdata(
    seed_dir: Path = SEED_DIR,
    batch_size: int = 25,
    limit: int = 0,
    refresh: bool = False,
) -> dict[str, object]:
    ensure_data_dirs()
    if batch_size <= 0:
        raise ValueError("batch_size must be greater than 0")
    linkedin_rows = _read_csv(seed_dir / "linkedin_profiles.csv")
    raw_path = INTERIM_DIR / "brightdata_linkedin_profiles.jsonl"
    normalized_path = AUDIT_DIR / "brightdata_profile_decisions.csv"

    existing_by_url: dict[str, dict[str, object]] = {}
    if not refresh:
        for item in _iter_brightdata_cached_records(raw_path):
            if item.get("input_url"):
                existing_by_url[str(item["input_url"])] = item

    candidates = [
        row
        for row in linkedin_rows
        if row.get("linkedin_url")
        and row.get("linkedin_url") != "N/A"
        and is_linkedin_profile_url(row.get("linkedin_url", ""))
        and row.get("linkedin_url") not in existing_by_url
    ]
    if limit > 0:
        candidates = candidates[:limit]
    total_batches = (len(candidates) + batch_size - 1) // batch_size if candidates else 0
    print(
        f"Bright Data enrichment: {len(candidates)} profiles pending, "
        f"batch_size={batch_size}, refresh={refresh}",
        flush=True,
    )
    client = BrightDataLinkedInClient()
    fetched: list[dict[str, object]] = []
    fetched_count = 0
    for offset in range(0, len(candidates), batch_size):
        batch = candidates[offset : offset + batch_size]
        urls = [row["linkedin_url"] for row in batch]
        batch_number = (offset // batch_size) + 1
        print(f"Bright Data batch {batch_number}/{total_batches}: requesting {len(urls)} profiles", flush=True)
        payload = client.scrape_profiles(urls)
        if isinstance(payload, list):
            records = payload
        elif isinstance(payload, dict) and isinstance(payload.get("data"), list):
            records = payload["data"]
        elif isinstance(payload, dict) and isinstance(payload.get("result"), list):
            records = payload["result"]
        elif isinstance(payload, dict) and isinstance(payload.get("result"), dict):
            records = [payload["result"]]
        elif isinstance(payload, dict):
            records = [payload]
        else:
            records = []
        if isinstance(records, dict):
            records = [records]
        batch_fetched = 0
        for index, record in enumerate(records):
            if isinstance(record, dict):
                input_url = _record_input_url(record, urls[index] if index < len(urls) else "")
                fetched.append({"input_url": input_url, "fetched_at": utc_now_iso(), "record": record})
                batch_fetched += 1
        if fetched:
            fetched_count += len(fetched)
            with raw_path.open("a", encoding="utf-8") as handle:
                for item in fetched:
                    handle.write(json.dumps(item, ensure_ascii=False) + "\n")
            for item in fetched:
                if item.get("input_url"):
                    existing_by_url[str(item["input_url"])] = item
            fetched = []
        print(
            f"Bright Data batch {batch_number}/{total_batches}: received {batch_fetched} records, "
            f"total_fetched={fetched_count}",
            flush=True,
        )

    appended_records = [item for item in existing_by_url.values() if "record" in item]
    if appended_records:
        raw_path.write_text("", encoding="utf-8")
        with raw_path.open("a", encoding="utf-8") as handle:
            for item in appended_records:
                handle.write(json.dumps(item, ensure_ascii=False) + "\n")

    all_records = list(existing_by_url.values())
    normalized_rows: list[dict[str, str]] = []
    for item in all_records:
        record = item.get("record", {})
        if isinstance(record, dict):
            normalized = normalize_brightdata_record(record, clean_text(item.get("input_url")))
            normalized["fetched_at"] = clean_text(item.get("fetched_at"))
            normalized_rows.append(normalized)
    if normalized_rows:
        _write_csv(normalized_path, normalized_rows, list(normalized_rows[0].keys()))
    return {"pending": len(candidates), "fetched": fetched_count, "raw": str(raw_path), "normalized": str(normalized_path)}


def _write_enrichlayer_progress(progress_path: Path, payload: dict[str, object]) -> None:
    progress_path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")


def _enrichlayer_resume_command(
    limit: int,
    fetched: int,
    delay_sec: float,
    max_retries: int,
    retry_after_sec: float,
) -> str:
    resume_limit = max(limit - fetched, 1) if limit > 0 else 0
    return (
        "python -m schwarzman_network.cli enrich-enrichlayer "
        f"--limit {resume_limit} --delay-sec {delay_sec:g} "
        f"--max-retries {max_retries} --retry-after-sec {retry_after_sec:g}"
    )


def enrich_enrichlayer(
    seed_dir: Path = SEED_DIR,
    limit: int = 200,
    refresh: bool = False,
    delay_sec: float = 0.0,
    max_retries: int = 1,
    retry_after_sec: float = 65.0,
) -> dict[str, object]:
    ensure_data_dirs()
    linkedin_rows = _read_csv(seed_dir / "linkedin_profiles.csv")
    normalized_path = AUDIT_DIR / "enrichlayer_profile_decisions.csv"
    progress_path = AUDIT_DIR / "enrichlayer_progress.json"
    existing_rows = _read_csv(normalized_path)
    existing_by_url = {
        normalize_linkedin_url(row.get("input_url", "")): row
        for row in existing_rows
        if normalize_linkedin_url(row.get("input_url", ""))
    }
    completed_urls = {
        url
        for url, row in existing_by_url.items()
        if row.get("enrichlayer_status") != "error"
    }
    candidates: list[tuple[int, dict[str, str], str]] = []
    for index, row in enumerate(linkedin_rows):
        linkedin_url = normalize_linkedin_url(row.get("linkedin_url", ""))
        if not linkedin_url or not is_linkedin_profile_url(linkedin_url):
            continue
        if not refresh and linkedin_url in completed_urls:
            continue
        candidates.append((index, row, linkedin_url))
    selected = candidates[:limit] if limit > 0 else candidates

    client = EnrichLayerClient()
    attempted = 0
    fetched = 0
    errors = 0
    stopped_reason = "nothing_to_do" if not selected else "limit_reached"
    last_row: dict[str, str] | None = None
    last_seed_index: int | None = None

    for selected_index, (seed_index, row, linkedin_url) in enumerate(selected, start=1):
        print(f"Enrichlayer {selected_index}/{len(selected)}: {_console_safe(linkedin_url)}", flush=True)
        fetched_at = utc_now_iso()
        normalized: dict[str, str] | None = None
        for attempt in range(max_retries + 1):
            try:
                record = client.fetch_profile(linkedin_url)
                normalized = normalize_enrichlayer_record(record, linkedin_url, fetched_at=fetched_at)
                fetched += 1
                break
            except EnrichLayerError as error:
                if error.status_code == 429 and attempt < max_retries:
                    print(f"Enrichlayer rate limited; sleeping {retry_after_sec:g}s before retry", flush=True)
                    time.sleep(retry_after_sec)
                    continue
                normalized = normalize_enrichlayer_record({}, linkedin_url, fetched_at=fetched_at, error=str(error))
                if error.status_code == 404:
                    normalized["enrichlayer_status"] = "not_found"
                errors += 1
                if error.status_code in {401, 402, 403, 429}:
                    stopped_reason = f"api_error_{error.status_code}"
                break
        if normalized is None:
            normalized = normalize_enrichlayer_record({}, linkedin_url, fetched_at=fetched_at, error="unknown_error")
            errors += 1
        existing_by_url[linkedin_url] = normalized
        attempted += 1
        last_row = row
        last_seed_index = seed_index
        _write_csv(normalized_path, list(existing_by_url.values()), ENRICHLAYER_FIELDNAMES)

        next_item = selected[selected_index] if selected_index < len(selected) else None
        progress = {
            "attempted_this_run": attempted,
            "fetched_this_run": fetched,
            "errors_this_run": errors,
            "requested_limit": limit,
            "refresh": refresh,
            "candidate_count_before_limit": len(candidates),
            "selected_count": len(selected),
            "last_seed_row_number": (last_seed_index + 2) if last_seed_index is not None else "",
            "last_scholar_id": last_row.get("scholar_id", "") if last_row else "",
            "last_scholar_name": last_row.get("scholar_name", "") if last_row else "",
            "last_linkedin_url": linkedin_url,
            "next_seed_row_number": (next_item[0] + 2) if next_item else "",
            "next_scholar_id": next_item[1].get("scholar_id", "") if next_item else "",
            "next_scholar_name": next_item[1].get("scholar_name", "") if next_item else "",
            "next_linkedin_url": next_item[2] if next_item else "",
            "stopped_reason": stopped_reason if stopped_reason.startswith("api_error_") else "",
            "updated_at": utc_now_iso(),
            "resume_command": _enrichlayer_resume_command(limit, fetched, delay_sec, max_retries, retry_after_sec),
        }
        _write_enrichlayer_progress(progress_path, progress)
        if stopped_reason.startswith("api_error_"):
            break
        if delay_sec > 0 and selected_index < len(selected):
            time.sleep(delay_sec)
    else:
        if selected and limit > 0 and len(candidates) > len(selected):
            stopped_reason = "limit_reached"
        elif selected:
            stopped_reason = "completed_all_pending"

    if stopped_reason.startswith("api_error_") and last_seed_index is not None and last_row:
        next_candidate = (
            last_seed_index,
            last_row,
            normalize_linkedin_url(last_row.get("linkedin_url", "")),
        )
    else:
        next_candidate = selected[attempted] if attempted < len(selected) else None
    progress = {
        "attempted_this_run": attempted,
        "fetched_this_run": fetched,
        "errors_this_run": errors,
        "requested_limit": limit,
        "refresh": refresh,
        "candidate_count_before_limit": len(candidates),
        "selected_count": len(selected),
        "last_seed_row_number": (last_seed_index + 2) if last_seed_index is not None else "",
        "last_scholar_id": last_row.get("scholar_id", "") if last_row else "",
        "last_scholar_name": last_row.get("scholar_name", "") if last_row else "",
        "last_linkedin_url": last_row.get("linkedin_url", "") if last_row else "",
        "next_seed_row_number": (next_candidate[0] + 2) if next_candidate else "",
        "next_scholar_id": next_candidate[1].get("scholar_id", "") if next_candidate else "",
        "next_scholar_name": next_candidate[1].get("scholar_name", "") if next_candidate else "",
        "next_linkedin_url": next_candidate[2] if next_candidate else "",
        "stopped_reason": stopped_reason,
        "updated_at": utc_now_iso(),
        "resume_command": _enrichlayer_resume_command(limit, fetched, delay_sec, max_retries, retry_after_sec),
    }
    _write_enrichlayer_progress(progress_path, progress)
    return {
        "pending_before_limit": len(candidates),
        "attempted": attempted,
        "fetched": fetched,
        "errors": errors,
        "stopped_reason": stopped_reason,
        "normalized": str(normalized_path),
        "progress": str(progress_path),
    }


def build_processed_profiles(seed_dir: Path = SEED_DIR, processed_path: Path | None = None, use_llm: bool = False) -> Path:
    ensure_data_dirs()
    scholars = _read_csv(seed_dir / "scholars.csv")
    linkedin_rows = {row["scholar_id"]: row for row in _read_csv(seed_dir / "linkedin_profiles.csv")}
    seed_observations = {row["scholar_id"]: row for row in _read_csv(seed_dir / "employment_observations.csv")}
    brightdata_by_url = {row["input_url"]: row for row in _read_csv(AUDIT_DIR / "brightdata_profile_decisions.csv")}
    enrichlayer_by_url = {
        row["input_url"]: row
        for row in _read_csv(AUDIT_DIR / "enrichlayer_profile_decisions.csv")
        if row.get("enrichlayer_status") == "ok"
    }
    out = processed_path or PROCESSED_DIR / "scholar_information.csv"

    rows: list[dict[str, str]] = []
    for scholar in scholars:
        scholar_id = scholar["scholar_id"]
        linkedin = linkedin_rows.get(scholar_id, {})
        linkedin_url = linkedin.get("linkedin_url") or "N/A"
        seed_observation = seed_observations.get(scholar_id, {})
        bright = brightdata_by_url.get(linkedin_url, {})
        enrich = enrichlayer_by_url.get(linkedin_url, {})
        current_company = _blank_if_na(
            enrich.get("enrichlayer_current_company")
            or bright.get("current_company_name")
            or seed_observation.get("current_company", "")
        )
        current_title = _blank_if_na(
            enrich.get("enrichlayer_current_job_title")
            or bright.get("position")
            or seed_observation.get("current_title", "")
        )
        profile_location = _blank_if_na(
            enrich.get("enrichlayer_profile_location")
            or bright.get("profile_location") or bright.get("location") or bright.get("profile_city") or bright.get("city")
            or seed_observation.get("current_location", "")
        )
        job_location = _blank_if_na(enrich.get("enrichlayer_current_job_location") or bright.get("job_location"))
        company = enrich_company(current_company, current_title, role_context=f"{current_title} {job_location}", use_llm=use_llm)
        rows.append(
            {
                "Scholar Name": scholar.get("scholar_name", ""),
                "Industry": company.industry,
                "Cohort": scholar.get("cohort", ""),
                "LinkedIn Address": linkedin_url,
                "Profile Location": profile_location,
                "Job Location": job_location,
                "Current Job Title": current_title,
                "Current Company": current_company,
                "Company Description": company.company_description,
                "Experience Count": enrich.get("enrichlayer_experience_count", ""),
                "Education Count": enrich.get("enrichlayer_education_count", ""),
                "Work History": enrich.get("enrichlayer_experience_json", ""),
                "Education": enrich.get("enrichlayer_education_json", ""),
                "Enrichment Source": "enrichlayer" if enrich else "brightdata" if bright else "",
                "Enrichment Status": enrich.get("enrichlayer_status", "") or bright.get("record_status", ""),
                "Country": scholar.get("country", ""),
                "Confidence": "enrichlayer" if enrich else "brightdata" if bright else seed_observation.get("confidence", ""),
                "Last Updated": enrich.get("enrichlayer_fetched_at") or bright.get("fetched_at") or seed_observation.get("observed_at", ""),
                "Source URLs": linkedin_url if linkedin_url != "N/A" else scholar.get("official_url", ""),
            }
        )

    headers = [
        "Scholar Name",
        "Industry",
        "Cohort",
        "LinkedIn Address",
        "Profile Location",
        "Job Location",
        "Current Job Title",
        "Current Company",
        "Company Description",
        "Experience Count",
        "Education Count",
        "Work History",
        "Education",
        "Enrichment Source",
        "Enrichment Status",
        "Country",
        "Confidence",
        "Last Updated",
        "Source URLs",
    ]
    _write_csv(out, rows, headers)
    return out


def build_db_and_exports(processed_path: Path | None = None) -> dict[str, object]:
    db_path = build_database(PUBLIC_DIR / "schwarzman_network.sqlite", processed_path=processed_path)
    exports = export_public(db_path, PUBLIC_DIR)
    return {"database": str(db_path), **{key: str(value) for key, value in exports.items()}}
